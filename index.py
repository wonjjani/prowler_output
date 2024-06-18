import pandas as pd
import json
from openpyxl import load_workbook
# 셀 너비 자동 조정
from openpyxl.utils import get_column_letter
from deep_translator import GoogleTranslator
import pyarrow

# JSON 파일 읽기
json_file_path = "C:/Users/eggra/Downloads/prowler_aws/prowler-output-553153918398-20240610071217.ocsf.json"
with open(json_file_path, 'r') as file:
    data = json.load(file)

# 데이터프레임으로 변환
df = pd.json_normalize(data)

# 컬럼 이름 한국어로 변경
df.columns = [
    "심각도 ID", "심각도", "상태", "상태 코드", "상세 상태", "상태 ID", "활동 이름", "활동 ID", 
    "리소스", "카테고리 이름", "카테고리 UID", "클래스 이름", "클래스 UID", "이벤트 시간", 
    "위험 세부 정보", "타입 UID", "타입 이름", "이벤트 코드", "제품 이름", "제품 공급자 이름", 
    "제품 버전", "버전", "검사 타입", "관련 URL", "카테고리", "의존", "관련", "메모", 
    "CIS-3.0 준수", "CIS-2.0 준수", "CIS-1.4 준수", "CIS-1.5 준수", "AWS 계정 보안 온보딩 준수", 
    "생성 시간", "설명", "제품 UID", "제목", "UID", "계정 이름", "계정 타입", "계정 타입 ID", 
    "계정 UID", "계정 라벨", "조직 이름", "조직 UID", "클라우드 제공자", "클라우드 리전", 
    "조치 설명", "조치 참조", "ENS-RD2022 준수", "AWS 잘 설계된 프레임워크 보안 기둥 준수", 
    "AWS 기본 보안 모범 사례 준수", "FedRamp 중간 개정 4 준수", "NIST-CSF 1.1 준수", 
    "ISO27001-2013 준수", "GDPR 준수", "RBI 사이버 보안 프레임워크 준수", "FFIEC 준수", 
    "NIST-800-53 개정 4 준수", "GxP 21 CFR 파트 11 준수", "CISA 준수", "GxP EU 부록 11 준수", 
    "HIPAA 준수", "NIST-800-171 개정 2 준수", "NIST-800-53 개정 5 준수", "SOC2 준수", 
    "PCI 3.2.1 준수", "FedRAMP 낮은 개정 4 준수", "MITRE ATT&CK 준수", "AWS 기본 기술 리뷰 준수", 
    "AWS 감사 관리자 컨트롤 타워 가드레일 준수"
]

# # "상태 코드"가 "PASS"인 행을 제거
# df = df[df["상태 코드"] != "PASS"]

# '상세 상태' 번역
translator = GoogleTranslator(source='en', target='ko')
df['상세 상태'] = df["상세 상태"].apply(lambda x: translator.translate(x) if pd.notnull(x) else x)
df['위험 세부 정보'] = df["위험 세부 정보"].apply(lambda x: translator.translate(x) if pd.notnull(x) else x)
df['설명'] = df["설명"].apply(lambda x: translator.translate(x) if pd.notnull(x) else x)
df['조치 설명'] = df["조치 설명"].apply(lambda x: translator.translate(x) if pd.notnull(x) else x)

# 제거할 컬럼 목록
columns_to_remove = [
    "제품 이름", "제품 공급자 이름", "제품 버전", "버전", "카테고리", "의존", "관련", "메모", 
    "CIS-2.0 준수", "CIS-1.4 준수", "CIS-1.5 준수", "AWS 계정 보안 온보딩 준수", "생성 시간", 
    "제품 UID", "UID", "계정 이름", "계정 타입", "계정 타입 ID", "계정 UID", "계정 라벨", 
    "조직 이름", "조직 UID", "클라우드 제공자", "ENS-RD2022 준수", 
    "AWS 잘 설계된 프레임워크 보안 기둥 준수", "AWS 기본 보안 모범 사례 준수", 
    "FedRamp 중간 개정 4 준수", "NIST-CSF 1.1 준수", "ISO27001-2013 준수", "GDPR 준수", 
    "RBI 사이버 보안 프레임워크 준수", "FFIEC 준수", "NIST-800-53 개정 4 준수", 
    "GxP 21 CFR 파트 11 준수", "CISA 준수", "GxP EU 부록 11 준수", "HIPAA 준수", 
    "NIST-800-171 개정 2 준수", "NIST-800-53 개정 5 준수", "SOC2 준수", "PCI 3.2.1 준수", 
    "FedRAMP 낮은 개정 4 준수", "MITRE ATT&CK 준수", "AWS 기본 기술 리뷰 준수", 
    "AWS 감사 관리자 컨트롤 타워 가드레일 준수", "제품 UID", "타입 UID", "타입 이름", "이벤트 시간",
    "카테고리 이름", "카테고리 UID", "클래스 이름", "클래스 UID", "상태 ID", "활동 이름", "활동 ID",
    "제목"
]

# 컬럼 제거
df = df.drop(columns=columns_to_remove, errors='ignore')

# 엑셀 파일로 저장
output_file_path = 'prowler_hahaha.xlsx'
df.to_excel(output_file_path, index=False)

wb = load_workbook(output_file_path)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 4)
    ws.column_dimensions[column].width = adjusted_width

wb.save(output_file_path)

print(f"엑셀 파일이 성공적으로 저장되었습니다: {output_file_path}")
